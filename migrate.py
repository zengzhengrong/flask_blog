import psycopg2
import time
import datetime
import os
from sqlalchemy.exc import IntegrityError,InvalidRequestError
from openpyxl import Workbook
from openpyxl import load_workbook
from flask_blog import create_app,db
from flask_blog.models import User,Post,Account

app = create_app()
# app.app_context().push()
def export_excel():
    conn = psycopg2.connect(database='flaskdb',user='postgres',password='5432',host='localhost')
    cursor = conn.cursor()
    wb = Workbook()
    tables_name = ['user','post','account']
    for name in tables_name:
        if name == 'user':
            exec(f'ws_user = wb.active')
            exec(f'ws_user.title = "{name} data"')
        else:
            exec(f'ws_{name} = wb.create_sheet(title="{name} data")')
        cursor.execute(f'select * from public.{name}')
        fields = cursor.description
        tb = cursor.fetchall()
        fields_name = [ field[0] for field in fields]
        exec(f'ws_{name}.append({fields_name})')
        exec(f"""for data in {tb}:
                ws_{name}.append(data)""")
    # 存放文件
    os.getcwd()
    file_dir = os.getcwd()+'/migrations/'
    if not os.path.exists(file_dir):
        os.mkdir(file_dir)
    ls = os.listdir(file_dir)
    next_number = 0 # 初始编号
    if ls:
        combine_tuple = tuple(zip(range(len(ls)),ls))
        # print(combine_tuple)
        next_number = combine_tuple[-1][0] + 1
    ex_file_name = f'migrations/mg{next_number}_blog.xlsx'
    wb.save(ex_file_name)
    print(f'数据库内容已迁移excel表:{ex_file_name}')
def make_migration():
    with app.app_context():
        db.drop_all()
        tables_name = ['user','post','account']
        for name in tables_name:
            conn = psycopg2.connect(database='flaskdb',user='postgres',password='5432',host='localhost')
            cursor = conn.cursor()
            try:
                cursor.execute(f'select * from public.{name}')
                conn.close()
            except Exception as e: # 需要将连接写入循环，因查询第一次后会触发连接断开
                if f'错误:  关系 "public.{name}" 不存在' in str(e):
                    print(f'确认删除表{name}')
                    conn.close()
                    # print(dir(cursor))
                else:
                    print(e)
        print('清空数据库成功')
        print('准备重新初始化数据库')
        db.create_all()
        print('初始化数据库成功')
    # 获取上次excel文件
    os.getcwd()
    file_dir = os.getcwd()+'/migrations/'
    ls = os.listdir(file_dir)
    next_number = 0 # 初始编号
    if ls:
        combine_tuple = tuple(zip(range(len(ls)),ls))
        print(combine_tuple)
        next_number = combine_tuple[-1][0]
    ex_file_name = f'migrations/mg{next_number}_blog.xlsx'
    wb = load_workbook(filename = ex_file_name)
    
    with psycopg2.connect(database='flaskdb',user='postgres',password='5432',host='localhost') as conn:
        cursor = conn.cursor()
        for name in tables_name:
            sheet = wb[f'{name} data']
            rows = sheet.rows
            # print(dir(rows))
            print('-'*12)
            model_name = name.title() # model名
            same_fields = None # 字段相同但是位置不一样时这个变量用来代替real_fields
            tag_delete = None # 字段减少标志重置
            tag_add = None # 字段增加标志重置
            diff_fields = None # 子循环完毕重置
            diff_fields_index = [] # 索引位列表，用于删除或插入line的数据
            for row in rows:
                cursor.execute(f'select * from public.{name}')
                fields = cursor.description
                # print(fields)
                fields_name = [ field[0] for field in fields]
                line = [col.value for col in row]
                real_fields = fields_name
                orig_fields = line
                # print(real_fields)
                # print(orig_fields)
                if fields_name == orig_fields: # 字段没有改变
                    print(f'The {name} table fields no change')
                    continue
                elif len(fields_name) < len(orig_fields) and tag_delete is not True: # 字段减少
                    diff_fields = [ field for field in orig_fields if field not in fields_name]
                    print(f'The {name} table fields have been changed ,{diff_fields} are removed')
                    tag_delete = True
                    # 字段减少导致的数据错位，需要删除对应字段的数据
                    if diff_fields:
                        for field in diff_fields:
                            index = orig_fields.index(field)
                            diff_fields_index.append(index)
                    # print(diff_fields_index)
                    continue
                elif len(fields_name) > len(orig_fields) and tag_add is not True: # 字段增加
                    diff_fields = [ field for field in real_fields if field not in orig_fields]
                    print(f'The {name} table fields have been changed ,{diff_fields} are added')
                    tag_add = True
                    # 字段增加导致的数据错位，需要插入对应字段的数据，默认插入None，字段要在数据库中允许为空
                    if diff_fields:
                        for field in diff_fields:
                            index = real_fields.index(field)
                            diff_fields_index.append(index)
                    continue            
                elif fields_name != orig_fields and orig_fields[0] == 'id': 
                    # print(fields_name == line)
                    diff_fields = [ field for field in orig_fields if field not in fields_name]
                    # print(diff_fields)
                    if not diff_fields and len(fields_name) == len(orig_fields): # 字段相同但是位置不一样
                        print(f'The {name} table fields are same before , but locations are difference')
                        same_fields = orig_fields
                        # print(same_fields)
                    if diff_fields:
                        raise AttributeError(f'{diff_fields} fields not found')
                    continue
                if same_fields:
                    import_db(name,same_fields,line)
                if diff_fields_index:
                    # print(diff_fields_index)
                    diff_index = 0
                    for index in diff_fields_index:
                        if tag_delete:
                            _data = line.pop(index-diff_index)
                            print(f'因{diff_fields[diff_index]}移除，{_data}将不被导入数据库')                            
                        if tag_add:
                            line.insert(index,None)
                            print(f'字段{diff_fields[diff_index]}增加，开始导入默认数据')
                            
                        diff_index += 1
                    dict_data = dict(zip(real_fields,line))
                    if tag_add:
                        list_data = list(zip(real_fields,line))
                        rm_index = 0
                        for index in diff_fields_index:
                            rm_data = list_data.pop(index-rm_index)
                            rm_index += 1 
                        print(f'该行导入默认数据完成')
                        dict_data = dict(list_data)
                    import_db(model_name,dict_data)
                    print('-'*8)
                if (not tag_add and not tag_delete) and not same_fields:
                    dict_data = dict(zip(real_fields,line))
                    import_db(model_name,dict_data)
    
def import_db(model_name,data):
    # print(data.items())
    with app.app_context():
        model_object = eval(f'{model_name}(**{data})')
        db.session.add(model_object)
        db.session.commit()
        print(f'导入一个{model_name}模型成功')
def test_db():
    tables_name = ['user','post','account']
    print('开始测试')
    for name in tables_name:
        model_name = name.title()
        with app.app_context():
            count = eval(f'len({model_name}.query.all())')
            print(f'准备测试{model_name}模型，测试数量为{count}')
            for i in range(count+1):
                with app.app_context(): # 重点:这里app上下文必须要在循环内，app.app_context().push()方式无效
                    try:
                        model = None
                        if model_name == 'User':
                            model = User(username='test',email='test@163.com',password='test123')
                        if model_name == 'Post':
                            test = User.query.filter_by(username='test').first()
                            model = Post(title='test',content='test'*3,author=test)
                        if model_name == 'Account':
                            test = User.query.filter_by(username='test').first()
                            model = Account()
                            model.user_id = test.id
                        if model is not None:
                            db.session.add(model)
                            db.session.commit()                           
                    except Exception as e:
                        time.sleep(0.01)
                        # print(e)
                        print(f'{model_name}模型执行第{i}次')
def drop_test_model():
    with app.app_context():
        test_user = User.query.filter_by(username='test').first()
        test_post = Post.query.filter_by(title='test').first()
        test_account = Account.query.filter_by(user=test_user).first()
        db.session.delete(test_post)
        db.session.delete(test_account)
        db.session.delete(test_user)
        db.session.commit()
        print('测试完成')
if __name__ == "__main__":
    export_excel()
    make_migration()
    test_db()
    drop_test_model()